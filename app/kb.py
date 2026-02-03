from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import select

from app.config import settings
from app.db import SessionLocal
from app.llm import embed_texts
from app.models import KBEntry


@dataclass(frozen=True)
class RetrievedKB:
    score: float
    category: str | None
    unit: str | None
    scope: str | None
    description: str | None
    answer: str


class KBStore:
    def __init__(self) -> None:
        self._property_registry: dict[str, dict[str, str]] = {}
        self._registry_rows: list[dict[str, str]] = []
        self._registry_by_id: dict[str, dict[str, str]] = {}
        self._registry_by_name: dict[str, dict[str, str]] = {}
        self._registry_sheet_name: str | None = None
        self._kb_sheet_name: str | None = None
        self._registry_key_field: str | None = None

    @property
    def property_registry(self) -> dict[str, dict[str, str]]:
        return self._property_registry

    def load_from_excel(self, excel_path: str) -> None:
        path = Path(excel_path)
        if not path.exists():
            return

        wb = load_workbook(filename=str(path), data_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            return

        registry_sheet_name = _pick_sheet_name(sheet_names, "Strutture", default=sheet_names[0])
        kb_sheet_name = _pick_sheet_name(
            sheet_names,
            "Knowledge base",
            default=sheet_names[1] if len(sheet_names) > 1 else sheet_names[0],
        )
        self._registry_sheet_name = registry_sheet_name
        self._kb_sheet_name = kb_sheet_name

        kb_sheet = wb[kb_sheet_name]
        headers = self._read_headers(kb_sheet)
        idx, _debug = self._build_header_index(headers)
        rows = list(self._iter_kb_rows(kb_sheet, idx=idx))

        registry_sheet = wb[registry_sheet_name]
        self._registry_rows = self._read_registry_rows(registry_sheet)
        self._build_registry_indexes()

        if rows:
            self._sync_rows(rows)

    def inspect_excel(self, excel_path: str) -> dict:
        path = Path(excel_path)
        if not path.exists():
            return {"ok": False, "error": "File not found"}

        wb = load_workbook(filename=str(path), data_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            return {"ok": False, "error": "No sheets found"}

        registry_sheet_name = _pick_sheet_name(sheet_names, "Strutture", default=sheet_names[0])
        kb_sheet_name = _pick_sheet_name(
            sheet_names,
            "Knowledge base",
            default=sheet_names[1] if len(sheet_names) > 1 else sheet_names[0],
        )

        kb_sheet = wb[kb_sheet_name]
        headers = self._read_headers(kb_sheet)
        idx, debug = self._build_header_index(headers)
        rows = list(self._iter_kb_rows(kb_sheet, idx=idx))
        sample = rows[:3]

        registry_sheet = wb[registry_sheet_name]
        registry_rows = self._read_registry_rows(registry_sheet)
        registry_sample = registry_rows[:3]
        registry_key_field = self._detect_registry_name_field(registry_rows)

        return {
            "ok": True,
            "sheet_names": sheet_names,
            "kb_sheet_name": kb_sheet_name,
            "registry_sheet_name": registry_sheet_name,
            "registry_key_field": registry_key_field,
            "headers": headers,
            "header_map": debug,
            "row_count_valid": len(rows),
            "row_count_total": kb_sheet.max_row - 1 if kb_sheet.max_row else 0,
            "sample_rows": sample,
            "registry_row_count": len(registry_rows),
            "registry_sample_rows": registry_sample,
        }

    def retrieve(
        self,
        query: str,
        *,
        property_hint: str | None,
        top_k: int | None = None,
    ) -> list[RetrievedKB]:
        top_k = top_k or settings.kb_top_k

        # Load all entries once per request (KB tipicamente piccola).
        with SessionLocal() as db:
            entries = list(db.scalars(select(KBEntry)).all())

        if not entries:
            return []

        query_vec = embed_texts([query])[0]

        candidates: list[tuple[float, KBEntry]] = []
        for entry in entries:
            if not self._matches_property(entry.unit, property_hint):
                continue
            emb = json.loads(entry.embedding_json)
            score = _cosine_similarity(query_vec, emb)
            candidates.append((float(score), entry))

        candidates.sort(key=lambda x: x[0], reverse=True)
        best = candidates[:top_k]
        return [
            RetrievedKB(
                score=score,
                category=e.category,
                unit=e.unit,
                scope=e.scope,
                description=e.description,
                answer=e.answer,
            )
            for score, e in best
        ]

    def _sync_rows(self, rows: list[dict[str, str | None]]) -> None:
        """
        Treat the Excel file as the source of truth:
        - delete KB rows that are no longer present
        - add new rows (and compute embeddings)
        """
        desired_hashes = {_hash_row(r) for r in rows}

        with SessionLocal() as db:
            existing = list(db.scalars(select(KBEntry)).all())
            existing_hashes = {e.row_hash for e in existing}

            # Remove stale rows
            stale = [e for e in existing if e.row_hash not in desired_hashes]
            for e in stale:
                db.delete(e)

            # Add missing rows
            missing = [r for r in rows if _hash_row(r) not in existing_hashes]
            if missing:
                texts = [_row_to_embedding_text(r) for r in missing]
                embeddings = embed_texts(texts)
                for row, emb in zip(missing, embeddings, strict=True):
                    db.add(
                        KBEntry(
                            row_hash=_hash_row(row),
                            category=row.get("Categoria"),
                            unit=row.get("Appartamento /stanza"),
                            scope=row.get("ambito"),
                            description=row.get("descrizione"),
                            answer=row.get("risposta") or "",
                            embedding_json=json.dumps(emb),
                        )
                    )
            db.commit()

    @staticmethod
    def _read_headers(sheet) -> list[str]:
        headers: list[str] = []
        for cell in sheet[1]:
            val = str(cell.value).strip() if cell.value is not None else ""
            headers.append(val)
        return headers

    @staticmethod
    def _iter_kb_rows(sheet, *, idx: dict[str, int]) -> Iterable[dict[str, str | None]]:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() for v in row):
                continue
            out: dict[str, str | None] = {}
            for key, i in idx.items():
                val = row[i] if i < len(row) else None
                sval = str(val).strip() if val is not None else None
                out[key] = sval if sval else None
            # require answer
            if not (out.get("risposta") and str(out["risposta"]).strip()):
                continue
            yield out

    @staticmethod
    def _read_registry_rows(sheet) -> list[dict[str, str]]:
        # Minimal generic loader: first row headers, each next row is a record.
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")

        registry_rows: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() for v in row):
                continue
            record: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = row[i] if i < len(row) else None
                if val is None:
                    continue
                sval = str(val).strip()
                if sval:
                    record[h] = sval
            if record:
                registry_rows.append(record)
        return registry_rows

    @staticmethod
    def _matches_property(unit_cell: str | None, property_hint: str | None) -> bool:
        if not unit_cell:
            return True
        unit_norm = unit_cell.strip().lower()
        if unit_norm in {"*", "all", "tutte", "tutti", "generale", "general"}:
            return True
        if not property_hint:
            # No hint: allow only "generale"
            return False
        hint = property_hint.strip().lower()
        return unit_norm == hint

    def resolve_property_name(self, property_id: str | None) -> tuple[str | None, dict[str, str] | None]:
        """
        Returns (property_name, registry_record) using:
        - property_id column if present
        - fallback to Nome (name) matching or direct property_id as name
        """
        if not property_id:
            return None, None
        pid = property_id.strip()
        if not pid:
            return None, None

        if pid in self._registry_by_id:
            record = self._registry_by_id[pid]
            name = record.get(self._registry_key_field or "Nome") or pid
            return name, record

        if pid in self._registry_by_name:
            record = self._registry_by_name[pid]
            return pid, record

        return pid, self._registry_by_name.get(pid)

    @staticmethod
    def _normalize_header(value: str) -> str:
        return (
            value.strip()
            .lower()
            .replace(" ", "")
            .replace("/", "")
            .replace("\\", "")
            .replace("-", "")
            .replace("_", "")
        )

    def _build_header_index(self, headers: list[str]) -> tuple[dict[str, int], dict[str, str]]:
        """
        Map flexible header names to canonical keys.
        Returns (idx, debug_map) where debug_map is canonical->original header.
        """
        canonical = {
            "categoria": "Categoria",
            "category": "Categoria",
            "appartamentostanza": "Appartamento /stanza",
            "appartamentostanze": "Appartamento /stanza",
            "appartamento": "Appartamento /stanza",
            "stanza": "Appartamento /stanza",
            "camera": "Appartamento /stanza",
            "struttura": "Appartamento /stanza",
            "property": "Appartamento /stanza",
            "ambito": "ambito",
            "scope": "ambito",
            "descrizione": "descrizione",
            "description": "descrizione",
            "risposta": "risposta",
            "answer": "risposta",
            "response": "risposta",
        }

        idx: dict[str, int] = {}
        debug: dict[str, str] = {}
        for i, h in enumerate(headers):
            norm = self._normalize_header(h)
            if norm in canonical:
                key = canonical[norm]
                if key not in idx:
                    idx[key] = i
                    debug[key] = h

        # If "risposta" is missing, fall back by position (expected order)
        # to avoid silently mis-mapping columns.
        if "risposta" not in idx and len(headers) >= 5:
            idx = {
                "Categoria": 0,
                "Appartamento /stanza": 1,
                "ambito": 2,
                "descrizione": 3,
                "risposta": 4,
            }
            debug = {k: headers[i] if i < len(headers) else "" for k, i in idx.items()}

        return idx, debug

    @staticmethod
    def _detect_registry_name_field(rows: list[dict[str, str]]) -> str | None:
        if not rows:
            return None
        # Prefer exact "Nome", else case-insensitive match
        for key in rows[0].keys():
            if key == "Nome":
                return "Nome"
        for key in rows[0].keys():
            if key.strip().lower() == "nome":
                return key
        return None

    def _build_registry_indexes(self) -> None:
        self._registry_by_id = {}
        self._registry_by_name = {}
        self._property_registry = {}

        if not self._registry_rows:
            return

        name_field = self._detect_registry_name_field(self._registry_rows) or "Nome"
        self._registry_key_field = name_field

        # detect property_id-like field
        property_id_field = None
        for key in self._registry_rows[0].keys():
            norm = self._normalize_header(key)
            if norm in {"propertyid", "idstruttura", "codicestruttura", "idproperty"}:
                property_id_field = key
                break

        for row in self._registry_rows:
            name_val = row.get(name_field) or row.get("Nome")
            if name_val:
                self._registry_by_name[name_val] = row
                self._property_registry[name_val] = row
            if property_id_field:
                pid = row.get(property_id_field)
                if pid:
                    self._registry_by_id[pid] = row


def _pick_sheet_name(sheet_names: list[str], preferred: str, *, default: str) -> str:
    for name in sheet_names:
        if name.strip().lower() == preferred.strip().lower():
            return name
    return default


def _row_to_embedding_text(row: dict[str, str | None]) -> str:
    parts = []
    for k in ("Categoria", "Appartamento /stanza", "ambito", "descrizione", "risposta"):
        v = row.get(k)
        if v:
            parts.append(f"{k}: {v}")
    return "\n".join(parts)


def _hash_row(row: dict[str, str | None]) -> str:
    blob = json.dumps(row, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()


def _cosine_similarity(a: list[float], b: list[float]) -> float:
    if not a or not b:
        return 0.0
    dot = 0.0
    norm_a = 0.0
    norm_b = 0.0
    for x, y in zip(a, b):
        dot += float(x) * float(y)
        norm_a += float(x) * float(x)
        norm_b += float(y) * float(y)
    denom = math.sqrt(norm_a) * math.sqrt(norm_b)
    if denom == 0.0:
        return 0.0
    return dot / denom
